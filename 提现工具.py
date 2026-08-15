from __future__ import annotations

import json
import os
import shutil
import sys
import threading
import traceback
import urllib.error
import urllib.request
from collections import defaultdict
from datetime import date, datetime, timedelta
from pathlib import Path
from tkinter import BOTH, END, LEFT, RIGHT, X, Button, Frame, Label, messagebox, scrolledtext, Tk

from openpyxl import Workbook, load_workbook
from openpyxl.formatting.rule import FormulaRule
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter


APP_DIR = Path(sys.executable).resolve().parent if getattr(sys, "frozen", False) else Path(__file__).resolve().parent
INPUT_DIR = APP_DIR / "每周钱包流水"
RESULT_DIR = APP_DIR / "提现结果"
CONFIRM_PATH = APP_DIR / "提现确认登记表.xlsx"
CONFIG_PATH = APP_DIR / "飞书配置.json"
CACHE_DIR = APP_DIR / "数据缓存"
WALLET_CACHE_PATH = CACHE_DIR / "钱包流水.json"
API = "https://open.feishu.cn/open-apis"


def canonical_withdrawal_status(value):
    value = text(value)
    if value == "opt1cHQuDn":
        return "未提现"
    if value == "optq4ExGVS":
        return "已发起待到账"
    if value == "opt956381830":
        return "规则未配置"
    return value


def text(value):
    if isinstance(value, list):
        return ",".join(filter(None, (text(x) for x in value)))
    if isinstance(value, dict):
        return text(value.get("name") or value.get("text") or value.get("value") or value.get("id") or "")
    return "" if value is None else str(value).strip()


def amount(value):
    try:
        return abs(float(value or 0))
    except (TypeError, ValueError):
        return 0.0


def number(value, fallback=0.0):
    try:
        return float(value)
    except (TypeError, ValueError):
        return fallback


def excel_value(value):
    """Convert Feishu rich values into values that Excel can store safely."""
    if value is None or isinstance(value, (str, int, float, bool, date, datetime)):
        return value
    if isinstance(value, (dict, list)):
        readable = text(value)
        return readable if readable else json.dumps(value, ensure_ascii=False)
    return str(value)


def key_of(row):
    return f'{text(row.get("交易时间"))}||{text(row.get("交易流水号"))}'


def record_id_of(row):
    return text(row.get("交易流水号"))


def safe_name(value):
    value = text(value) or "未命名"
    return "".join("_" if c in '\\/:*?\"<>|' else c for c in value)[:80]


def compact_day(value):
    return text(value)[:10].replace("-", "")


def read_rows(path, sheet_name=None):
    wb = load_workbook(path, read_only=True, data_only=False)
    ws = wb[sheet_name] if sheet_name and sheet_name in wb.sheetnames else wb.worksheets[0]
    rows = list(ws.iter_rows(values_only=True))
    wb.close()
    if not rows:
        return []
    headers = [text(x) for x in rows[0]]
    return [dict(zip(headers, row)) for row in rows[1:] if any(x not in (None, "") for x in row)]


def style_sheet(ws, widths):
    dark = PatternFill("solid", fgColor="1F4E78")
    light = PatternFill("solid", fgColor="D9EAF7")
    thin = Side(style="thin", color="D9E2F3")
    for cell in ws[1]:
        cell.fill = light
        cell.font = Font(bold=True, color="17365D")
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = Border(bottom=thin)
    for i, width in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = width
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = ws.dimensions
    ws.sheet_view.showGridLines = False
    for row in ws.iter_rows(min_row=2):
        for cell in row:
            cell.border = Border(bottom=thin)
            cell.alignment = Alignment(vertical="center")


def append_rows(ws, headers, rows):
    ws.append(headers)
    for row in rows:
        ws.append([excel_value(row.get(h, "")) for h in headers])


class Feishu:
    def __init__(self, config):
        self.config = config
        self.app_token = config["app_token"]
        self._field_cache = {}
        response = self._http("POST", f"{API}/auth/v3/tenant_access_token/internal",
                              {"app_id": config["app_id"], "app_secret": config["app_secret"]}, token="")
        if response.get("code") != 0:
            raise RuntimeError(f'飞书登录失败：{response.get("msg")}')
        self.token = response["tenant_access_token"]

    @staticmethod
    def _http(method, url, body=None, token=""):
        data = json.dumps(body).encode("utf-8") if body is not None else None
        headers = {"Content-Type": "application/json; charset=utf-8"}
        if token:
            headers["Authorization"] = f"Bearer {token}"
        request = urllib.request.Request(url, data=data, headers=headers, method=method)
        try:
            with urllib.request.urlopen(request, timeout=60) as response:
                return json.loads(response.read().decode("utf-8"))
        except urllib.error.HTTPError as exc:
            detail = exc.read().decode("utf-8", errors="replace")
            raise RuntimeError(f"飞书网络请求失败：HTTP {exc.code} {detail}") from exc

    def request(self, method, path, body=None):
        payload = self._http(method, f"{API}{path}", body, self.token)
        if payload.get("code") != 0:
            raise RuntimeError(f'飞书接口失败：{payload.get("code")} {payload.get("msg")}')
        return payload.get("data", payload)

    def list_records(self, table_id):
        items, token = [], ""
        while True:
            suffix = f"&page_token={token}" if token else ""
            data = self.request("GET", f"/bitable/v1/apps/{self.app_token}/tables/{table_id}/records?page_size=500{suffix}")
            items.extend(data.get("items", []))
            if not data.get("has_more"):
                return items
            token = data.get("page_token", "")

    def fields(self, table_id):
        if table_id in self._field_cache:
            return self._field_cache[table_id]
        data = self.request("GET", f"/bitable/v1/apps/{self.app_token}/tables/{table_id}/fields?page_size=500")
        self._field_cache[table_id] = data.get("items", [])
        return self._field_cache[table_id]

    def writable_row(self, table_id, row):
        fields = {x["field_name"]: x for x in self.fields(table_id)}
        readonly = {19, 20, 1001, 1002, 1005}
        result = {}
        for name, value in row.items():
            field = fields.get(name)
            if not field or field.get("type") in readonly:
                continue
            if field.get("type") == 2:
                try: value = float(value)
                except (TypeError, ValueError): continue
            elif field.get("type") == 5:
                try: value = int(datetime.strptime(text(value)[:10], "%Y-%m-%d").timestamp() * 1000)
                except ValueError: continue
            elif value is None:
                value = ""
            result[name] = value
        return result

    def option_maps(self, table_id):
        """Resolve Feishu lookup/select option IDs to their visible labels."""
        maps = {}
        for field in self.fields(table_id):
            options = ((field.get("property") or {}).get("options")) or []
            if options:
                maps[field.get("field_name")] = {str(x.get("id")): text(x.get("name")) for x in options}
        target_cache = {}
        for field in self.fields(table_id):
            prop = field.get("property") or {}
            target_table = (prop.get("filter_info") or {}).get("target_table")
            target_field_id = prop.get("target_field")
            if not target_table or not target_field_id:
                continue
            if target_table not in target_cache:
                target_cache[target_table] = self.fields(target_table)
            target_field = next((x for x in target_cache[target_table] if x.get("field_id") == target_field_id), None)
            options = ((target_field or {}).get("property") or {}).get("options") or []
            if options:
                maps[field.get("field_name")] = {str(x.get("id")): text(x.get("name")) for x in options}
        return maps

    @staticmethod
    def visible_value(value, option_map=None):
        if value is None:
            return ""
        if isinstance(value, list):
            return ",".join(filter(None, (Feishu.visible_value(x, option_map) for x in value)))
        if isinstance(value, dict):
            return Feishu.visible_value(value.get("name") or value.get("text") or value.get("value") or value.get("id") or "", option_map)
        return (option_map or {}).get(str(value), value)

    def clear(self, table_id):
        ids = [x["record_id"] for x in self.list_records(table_id)]
        for i in range(0, len(ids), 500):
            self.request("POST", f"/bitable/v1/apps/{self.app_token}/tables/{table_id}/records/batch_delete", {"records": ids[i:i+500]})
        return len(ids)

    def create(self, table_id, rows):
        rows = [self.writable_row(table_id, row) for row in rows]
        for i in range(0, len(rows), 500):
            self.request("POST", f"/bitable/v1/apps/{self.app_token}/tables/{table_id}/records/batch_create",
                         {"records": [{"fields": row} for row in rows[i:i+500]]})

    def sync_by_key(self, table_id, rows, key):
        existing = {text(x.get("fields", {}).get(key)): x for x in self.list_records(table_id)}
        creates, updates = [], []
        for row in rows:
            old = existing.get(text(row.get(key)))
            row = self.writable_row(table_id, row)
            if old:
                updates.append({"record_id": old["record_id"], "fields": row})
            else:
                creates.append({"fields": row})
        for i in range(0, len(creates), 500):
            self.request("POST", f"/bitable/v1/apps/{self.app_token}/tables/{table_id}/records/batch_create", {"records": creates[i:i+500]})
        for i in range(0, len(updates), 500):
            self.request("POST", f"/bitable/v1/apps/{self.app_token}/tables/{table_id}/records/batch_update", {"records": updates[i:i+500]})

    def sync_wallet_ledger(self, table_id, new_rows, status_updates):
        existing_records = self.list_records(table_id)
        existing = {text(x.get("fields", {}).get("交易流水号")): x for x in existing_records}
        creates, updates = [], []
        for row in new_rows:
            transaction_id = text(row.get("交易流水号"))
            if not transaction_id or transaction_id in existing:
                continue
            creates.append({
                "fields": {
                    "交易流水号": transaction_id,
                    "店铺名称": text(row.get("店铺名称")),
                    "交易类型": text(row.get("交易类型")),
                    "币种": text(row.get("币种")),
                    "资金流向": text(row.get("资金流向")),
                    "交易金额": number(row.get("交易金额"), 0),
                    "交易状态": text(row.get("交易状态")),
                    "交易时间": self.writable_row(table_id, {"交易时间": row.get("交易时间")}).get("交易时间"),
                    "数据来源": text(row.get("数据来源")),
                    "导入日期": self.writable_row(table_id, {"导入日期": row.get("导入日期")}).get("导入日期"),
                    "提现状态": canonical_withdrawal_status(row.get("提现状态")),
                },
            })
        for row in status_updates:
            transaction_id = text(row.get("交易流水号"))
            old = existing.get(transaction_id)
            if not old:
                continue
            updates.append({
                "record_id": old["record_id"],
                "fields": {"提现状态": canonical_withdrawal_status(row.get("提现状态"))},
            })
        for i in range(0, len(creates), 500):
            self.request("POST", f"/bitable/v1/apps/{self.app_token}/tables/{table_id}/records/batch_create", {"records": creates[i:i+500]})
        for i in range(0, len(updates), 500):
            self.request("POST", f"/bitable/v1/apps/{self.app_token}/tables/{table_id}/records/batch_update", {"records": updates[i:i+500]})
        return len(existing_records), len(creates), len(updates)


def load_config():
    config = json.loads(CONFIG_PATH.read_text(encoding="utf-8-sig"))
    if config.get("app_id", "").startswith("请填写") or config.get("app_secret", "").startswith("请填写"):
        raise RuntimeError("请先打开“飞书配置.json”，填写 app_id 和 app_secret。")
    config.setdefault("tables", {}).setdefault("钱包流水", "tbl3AgMHFwFoX434")
    return config


def pull_feishu(feishu, tables):
    mapping_table = tables["店铺户主对应表"]
    option_maps = feishu.option_maps(mapping_table)
    mappings = [
        {name: feishu.visible_value(value, option_maps.get(name)) for name, value in x.get("fields", {}).items()}
        for x in feishu.list_records(mapping_table)
    ]
    receipts = [x.get("fields", {}) for x in feishu.list_records(tables["实际到账登记"])]
    wallet_option_maps = feishu.option_maps(tables["钱包流水"])
    wallet_rows = [
        {name: feishu.visible_value(value, wallet_option_maps.get(name)) for name, value in x.get("fields", {}).items()}
        for x in feishu.list_records(tables["钱包流水"])
    ]
    CACHE_DIR.mkdir(exist_ok=True)
    (CACHE_DIR / "店铺规则.json").write_text(json.dumps(mappings, ensure_ascii=False, indent=2), encoding="utf-8")
    (CACHE_DIR / "实际到账.json").write_text(json.dumps(receipts, ensure_ascii=False, indent=2), encoding="utf-8")
    WALLET_CACHE_PATH.write_text(json.dumps(wallet_rows, ensure_ascii=False, indent=2), encoding="utf-8")
    return mappings, receipts, wallet_rows


def load_state(wallet_rows):
    records = {}
    for row in wallet_rows:
        transaction_id = record_id_of(row)
        if not transaction_id:
            continue
        records[transaction_id] = {
            "id": transaction_id,
            "key": key_of(row),
            "transactionId": transaction_id,
            "store": text(row.get("店铺名称")) or "未命名店铺",
            "transactionType": text(row.get("交易类型")),
            "currency": text(row.get("币种")),
            "direction": text(row.get("资金流向")),
            "sourceStatus": text(row.get("交易状态")),
            "transactionTime": text(row.get("交易时间")),
            "originalAmount": number(row.get("交易金额"), 0),
            "absoluteAmount": amount(row.get("交易金额")),
            "withdrawalStatus": canonical_withdrawal_status(row.get("提现状态")) or "未提现",
            "withdrawalBatch": "",
            "confirmedAt": "",
            "sourceFile": text(row.get("数据来源")),
            "importedAt": text(row.get("导入日期")),
        }
    return {"version": 2, "records": records}
def run_process(first_run=False, logger=print):
    INPUT_DIR.mkdir(exist_ok=True)
    RESULT_DIR.mkdir(exist_ok=True)
    files = sorted(p for p in INPUT_DIR.glob("*.xlsx") if not p.name.startswith(("~$", ".~")))
    if not files:
        raise RuntimeError("每周钱包流水文件夹中没有 Excel 文件。")
    config = load_config()
    tables = config["tables"]
    logger("连接飞书并读取店铺规则、人工到账登记、钱包流水主表……")
    feishu = Feishu(config)
    mappings, receipts, wallet_rows = pull_feishu(feishu, tables)
    if first_run:
        logger("首次运行：清空飞书程序生成表……")
        for name in ("服务商提现批次", "店铺提现明细", "实时提现审核", "实时明细"):
            feishu.clear(tables[name])

    run_day = date.today().isoformat()
    out_dir = RESULT_DIR / run_day
    if first_run and out_dir.exists():
        shutil.rmtree(out_dir)
    out_dir.mkdir(parents=True, exist_ok=True)
    state = load_state([] if first_run else wallet_rows)
    seen = set(state["records"])
    imported = duplicate = 0
    new_wallet_rows = []
    for file in files:
        logger(f"读取：{file.name}")
        for row in read_rows(file):
            transaction_id = record_id_of(row)
            key = key_of(row)
            absolute_amount = amount(row.get("交易金额"))
            if not transaction_id or not text(row.get("交易时间")) or absolute_amount <= 0:
                continue
            if transaction_id in seen:
                duplicate += 1
                continue
            seen.add(transaction_id)
            imported += 1
            record = {
                "id": transaction_id, "key": key, "transactionId": transaction_id,
                "store": text(row.get("店铺名称")) or "未命名店铺",
                "transactionType": text(row.get("交易类型")), "currency": text(row.get("币种")),
                "direction": text(row.get("资金流向")), "sourceStatus": text(row.get("交易状态")),
                "transactionTime": text(row.get("交易时间")), "originalAmount": float(row.get("交易金额") or 0),
                "absoluteAmount": absolute_amount, "withdrawalStatus": "未提现",
                "withdrawalBatch": "", "confirmedAt": "", "sourceFile": file.name, "importedAt": run_day,
            }
            state["records"][transaction_id] = record
            new_wallet_rows.append({
                "交易流水号": record["transactionId"], "店铺名称": record["store"], "交易类型": record["transactionType"],
                "币种": record["currency"], "资金流向": record["direction"], "交易金额": record["originalAmount"],
                "交易状态": record["sourceStatus"], "交易时间": record["transactionTime"], "提现状态": record["withdrawalStatus"],
                "数据来源": record["sourceFile"], "导入日期": record["importedAt"],
            })

    rules = {text(x.get("店铺名称")): x for x in mappings if text(x.get("店铺名称"))}
    pending = defaultdict(list)
    for row in state["records"].values():
        if row.get("withdrawalStatus") == "未提现" and row.get("absoluteAmount", 0) > 0:
            pending[row["store"]].append(row)

    decisions = []
    realtime_rows = []
    status_updates = []
    for store in sorted({x["store"] for x in state["records"].values()}):
        rule = rules.get(store, {})
        mode = text(rule.get("提现模式")) or "未配置"
        rows = sorted(pending.get(store, []), key=lambda x: x["transactionTime"])
        total = sum(x["absoluteAmount"] for x in rows)
        if "实时" in mode:
            realtime_store_rows = sorted(
                [x for x in state["records"].values() if x["store"] == store and x.get("absoluteAmount", 0) > 0],
                key=lambda x: x["transactionTime"],
            )
            for idx, row in enumerate(realtime_store_rows, 1):
                previous_status = row.get("withdrawalStatus")
                row["withdrawalStatus"] = "次提人工处理"
                if previous_status != row["withdrawalStatus"]:
                    status_updates.append({"交易流水号": row["transactionId"], "提现状态": row["withdrawalStatus"]})
                realtime_rows.append({"店铺名称": store, "户主姓名": text(rule.get("户主姓名")), "服务商": text(rule.get("服务商")),
                                      "绝对值金额": row["absoluteAmount"], "交易时间": row["transactionTime"], "交易流水号": row["transactionId"],
                                      "原交易金额": row["originalAmount"], "区间顺序": idx, "数据来源": row["sourceFile"], "去重键": row["key"]})
            continue
        threshold = number(rule.get("提现门槛"), 30000)
        if "周四全部" in mode:
            should = total > 0
            reason = "运行时无门槛，全部进入提现" if should else "暂无未提现金额"
        elif "每周门槛" in mode:
            should = total >= threshold
            reason = f"运行时已达到门槛 {threshold:g}" if should else f"未达到提现门槛 {threshold:g}"
        else:
            should = False
            reason = "提现规则未配置"
            if total > 0:
                for row in rows:
                    previous_status = row["withdrawalStatus"]
                    row["withdrawalStatus"] = "规则未配置"
                    if previous_status != row["withdrawalStatus"]:
                        status_updates.append({"交易流水号": row["transactionId"], "提现状态": row["withdrawalStatus"]})
        decisions.append({"store": store, "rule": rule, "mode": mode, "rows": rows, "total": total, "should": should, "reason": reason})

    grouped = defaultdict(list)
    for d in decisions:
        if d["should"]:
            grouped[(text(d["rule"].get("服务商")) or "未配置服务商", d["mode"], "TWD")].append(d)
    existing_batches = [{name: feishu.visible_value(value) for name, value in x.get("fields", {}).items()} for x in feishu.list_records(tables["服务商提现批次"])]
    existing_details = [{name: feishu.visible_value(value) for name, value in x.get("fields", {}).items()} for x in feishu.list_records(tables["店铺提现明细"])]
    used_ids = {text(x.get("服务商批次")) for x in existing_batches}
    new_batches, new_details = [], []
    for (provider, mode, currency), items in grouped.items():
        base = f"{run_day}-{provider}-{mode}"
        batch_id, seq = base, 2
        while batch_id in used_ids:
            batch_id = f"{base}-{seq:02d}"
            seq += 1
        used_ids.add(batch_id)
        expected = sum(x["total"] for x in items)
        lag = int(number(items[0]["rule"].get("正常到账天数"), 0))
        expected_day = (date.fromisoformat(run_day) + timedelta(days=lag)).isoformat() if lag else ""
        new_batches.append({"服务商批次": batch_id, "处理日期": run_day, "服务商": provider, "提现模式": mode, "币种": currency,
                            "店铺数": len(items), "明细笔数": sum(len(x["rows"]) for x in items), "应提现金额": expected,
                            "发起提现日期": run_day, "账期": text(items[0]["rule"].get("账期")), "正常到账天数": lag,
                            "预计到账日": expected_day, "累计实际到账": 0, "未到账差额": expected, "到账比例": 0, "当前状态": "已发起待到账"})
        for d in items:
            store_batch = f'{batch_id}-{d["store"]}'
            for row in d["rows"]:
                previous_status = row["withdrawalStatus"]
                row["withdrawalStatus"] = "已发起待到账"
                row["withdrawalBatch"] = batch_id
                if previous_status != row["withdrawalStatus"]:
                    status_updates.append({"交易流水号": row["transactionId"], "提现状态": row["withdrawalStatus"]})
            d["batch"] = batch_id
            new_details.append({"服务商批次": batch_id, "店铺批次": store_batch, "处理日期": run_day, "服务商": provider,
                                "提现模式": mode, "店铺名称": d["store"], "卖家名称": text(d["rule"].get("卖家名称")),
                                "户主姓名": text(d["rule"].get("户主姓名")), "应提现金额": d["total"], "店铺状态": "已发起待到账"})
    merged_batches = {text(x.get("服务商批次")): x for x in existing_batches if text(x.get("服务商批次"))}
    for row in new_batches:
        merged_batches[text(row.get("服务商批次"))] = row
    merged_details = {f'{text(x.get("服务商批次"))}||{text(x.get("店铺批次"))}': x for x in existing_details if text(x.get("服务商批次")) or text(x.get("店铺批次"))}
    for row in new_details:
        merged_details[f'{text(row.get("服务商批次"))}||{text(row.get("店铺批次"))}'] = row
    all_batches = list(merged_batches.values())
    all_details = list(merged_details.values())
    build_outputs(out_dir, run_day, decisions, all_batches, receipts, all_details, imported, duplicate, new_batches)

    logger("同步飞书钱包流水主表、服务商批次和店铺明细……")
    existing_wallet, wallet_created, wallet_updated = feishu.sync_wallet_ledger(tables["钱包流水"], [
        {**row, "提现状态": next((x["提现状态"] for x in status_updates if x["交易流水号"] == row["交易流水号"]), row["提现状态"])}
        for row in new_wallet_rows
    ], status_updates)
    feishu.sync_by_key(tables["服务商提现批次"], all_batches, "服务商批次")
    feishu.sync_by_key(tables["店铺提现明细"], all_details, "店铺批次")
    existing_realtime = {text(x.get("fields", {}).get("去重键")) for x in feishu.list_records(tables["实时明细"])}
    feishu.create(tables["实时明细"], [x for x in realtime_rows if x["去重键"] not in existing_realtime])
    return {"导入": imported, "重复": duplicate, "批次": len(new_batches), "店铺明细": len(new_details),
            "实时明细": len(realtime_rows), "本次提现": sum(x["应提现金额"] for x in new_batches),
            "钱包流水已存": existing_wallet, "钱包流水新增": wallet_created, "钱包流水状态更新": wallet_updated, "结果目录": str(out_dir)}


def build_outputs(out_dir, run_day, decisions, batches, receipts, details, imported, duplicate, new_batches):
    wb = Workbook()
    ws = wb.active
    ws.title = "服务商提现批次"
    bh = ["服务商批次", "处理日期", "服务商", "提现模式", "币种", "店铺数", "明细笔数", "应提现金额", "发起提现日期", "账期", "正常到账天数", "预计到账日", "累计实际到账", "未到账差额", "到账比例", "当前状态"]
    append_rows(ws, bh, batches)
    style_sheet(ws, [30, 14, 14, 20, 10, 10, 12, 18, 15, 12, 15, 15, 18, 18, 14, 22])
    for c in (8, 13, 14):
        for cell in ws[get_column_letter(c)][1:]: cell.number_format = '#,##0.00'
    rws = wb.create_sheet("实际到账登记")
    rh = ["到账流水号", "服务商批次", "到账日期", "本次到账金额", "操作人", "银行凭证", "备注"]
    append_rows(rws, rh, receipts)
    style_sheet(rws, [20, 30, 15, 18, 14, 24, 34])
    dws = wb.create_sheet("店铺提现明细")
    dh = ["服务商批次", "店铺批次", "处理日期", "服务商", "提现模式", "店铺名称", "卖家名称", "户主姓名", "应提现金额", "店铺状态"]
    append_rows(dws, dh, details)
    style_sheet(dws, [30, 30, 14, 14, 20, 14, 24, 18, 18, 20])
    wb.save(CONFIRM_PATH)

    master = Workbook()
    ms = master.active
    ms.title = "本周汇总"
    headers = ["店铺名称", "卖家名称", "户主姓名", "服务商", "提现模式", "待处理笔数", "待处理金额", "系统状态", "判断原因", "本次应提现", "结转金额", "提现批次", "对账文件"]
    ms.append([f"店铺提现汇总（{run_day}）"])
    ms.append(["新增记录", imported, "重复记录", duplicate, "本次服务商批次", len(new_batches)])
    ms.append([])
    ms.append(headers)
    for d in decisions:
        owner = text(d["rule"].get("户主姓名"))
        start = min((x["transactionTime"][:10] for x in d["rows"]), default=run_day)
        end = max((x["transactionTime"][:10] for x in d["rows"]), default=run_day)
        filename = f"{safe_name(owner or d['store'])}_提现_{compact_day(start)}_{compact_day(end)}.xlsx" if d.get("batch") else ""
        ms.append([d["store"], text(d["rule"].get("卖家名称")), owner, text(d["rule"].get("服务商")), d["mode"], len(d["rows"]), d["total"],
                   "已发起待到账" if d["should"] else "未提现", d["reason"], d["total"] if d["should"] else 0,
                   0 if d["should"] else d["total"], d.get("batch", ""), filename])
        if d.get("batch"):
            daily = defaultdict(float)
            for row in d["rows"]: daily[row["transactionTime"][:10]] += row["absoluteAmount"]
            owb = Workbook(); ows = owb.active; ows.title = "Sheet1"
            oh = ["卖家账号", "交易时间", "户主姓名", "交易金额", "描述", "狀態", "提款到賬銀行"]
            ows.append(oh)
            for day, value in sorted(daily.items()):
                ows.append([text(d["rule"].get("卖家账号")), day, owner, value, "自動提款", "已轉入錢包", "银行"])
            ows.append(["", "", "", sum(daily.values()), "合计", "", ""])
            style_sheet(ows, [22, 16, 18, 16, 16, 18, 20])
            owb.save(out_dir / filename)
    style_sheet(ms, [14, 24, 18, 14, 20, 12, 16, 18, 30, 16, 16, 32, 40])
    ms.freeze_panes = "A5"
    master.save(out_dir / f"提现总表_{run_day}.xlsx")

    by_provider = defaultdict(list)
    for batch in new_batches:
        provider = batch["服务商"]
        for d in decisions:
            if d.get("batch") == batch["服务商批次"]:
                by_provider[provider].append((text(d["rule"].get("户主姓名")) or d["store"], d["total"]))
    for provider, rows in by_provider.items():
        lines = [f"{provider} 提现汇总 {run_day}", ""] + [f"{name}\t{value:,.0f}" for name, value in rows]
        lines += ["", f"户主数量：{len(rows)}", f"总金额：{sum(x[1] for x in rows):,.0f}"]
        (out_dir / f"{safe_name(provider)}_提现汇总_{compact_day(run_day)}.txt").write_text("\n".join(lines), encoding="utf-8-sig")


class App:
    def __init__(self):
        self.root = Tk()
        self.root.title("每周店铺提现工具")
        self.root.geometry("760x540")
        Label(self.root, text="每周店铺提现工具", font=("Microsoft YaHei UI", 20, "bold")).pack(pady=(18, 6))
        Label(self.root, text="把 BigSeller 钱包流水放进“每周钱包流水”文件夹，然后点击运行。", font=("Microsoft YaHei UI", 11)).pack(pady=4)
        bar = Frame(self.root); bar.pack(fill=X, padx=24, pady=14)
        Button(bar, text="打开流水文件夹", width=18, command=lambda: self.open_path(INPUT_DIR)).pack(side=LEFT, padx=5)
        Button(bar, text="正常每周运行", width=18, bg="#2E75B6", fg="white", command=lambda: self.start(False)).pack(side=LEFT, padx=5)
        Button(bar, text="首次运行（清空重建）", width=22, bg="#C65911", fg="white", command=lambda: self.confirm_first()).pack(side=LEFT, padx=5)
        Button(bar, text="打开结果", width=14, command=lambda: self.open_path(RESULT_DIR)).pack(side=RIGHT, padx=5)
        self.log = scrolledtext.ScrolledText(self.root, font=("Consolas", 10), height=22)
        self.log.pack(fill=BOTH, expand=True, padx=24, pady=(0, 18))
        INPUT_DIR.mkdir(exist_ok=True); RESULT_DIR.mkdir(exist_ok=True); CACHE_DIR.mkdir(exist_ok=True)

    def open_path(self, path):
        path.mkdir(parents=True, exist_ok=True)
        os.startfile(path)

    def write(self, value):
        self.root.after(0, lambda: (self.log.insert(END, value + "\n"), self.log.see(END)))

    def confirm_first(self):
        if messagebox.askyesno("确认首次运行", "将不继承历史台账和旧批次，并清空飞书程序生成表。\n店铺规则和人工到账登记不会删除。是否继续？"):
            self.start(True)

    def start(self, first):
        self.log.delete("1.0", END)
        threading.Thread(target=self.worker, args=(first,), daemon=True).start()

    def worker(self, first):
        try:
            result = run_process(first, self.write)
            self.write("\n处理完成：" + json.dumps(result, ensure_ascii=False, indent=2))
            self.root.after(0, lambda: messagebox.showinfo("完成", f'处理完成。\n本次提现：{result["本次提现"]:,.0f}\n结果已保存。'))
        except Exception as exc:
            CACHE_DIR.mkdir(exist_ok=True)
            (CACHE_DIR / "错误日志.txt").write_text(traceback.format_exc(), encoding="utf-8")
            self.write("处理失败：" + str(exc))
            self.root.after(0, lambda: messagebox.showerror("处理失败", f"{exc}\n\n详细信息已写入数据缓存\\错误日志.txt"))

    def run(self):
        self.root.mainloop()


if __name__ == "__main__":
    App().run()
